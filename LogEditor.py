import xlsxwriter
import datetime
import openpyxl
import sys
import os

class LogFile():

    def __init__(self, log_dir_name="PenTooLLogs"):
        self.current_dir = os.getcwd()
        self.log_dir_name = log_dir_name
        self.home_dir = os.path.expanduser("~")
        self.log_dir_path = os.path.join(self.home_dir, "Desktop", self.log_dir_name)
        self.date = datetime.datetime.today().strftime('%d-%m-%Y')
        self.log_file_name = "Log[{0}].xlsx".format(self.date)
        self.log_file_path = os.path.join(self.log_dir_path, self.log_file_name)
        # Aşağıdaki iki kod satırı dizin ve dosyayı oluşturmaktadır.
        self.create_log_directory()
        self.create_log_xlsx_file()

    def create_log_directory(self):

        print("---" * 37)
        try:
            os.mkdir(self.log_dir_path)
        except FileExistsError as error:
            print(error)
        else:
            print("Logların bulunacağı log dosyası '{0}' adresine başarılı bir " \
                  "şekilde oluşturuldu.".format(self.log_dir_path))
            print("---" * 37)

    def create_log_xlsx_file(self):

        try:
            if self.log_file_name not in os.listdir(self.log_dir_path):
                os.chdir(self.log_dir_path)
                # date = datetime.datetime.now().strftime("%d/%m/%Y")
                # date = str(date)
                # name = "Log-{0}-.xlsx".format(date)
                workbook = xlsxwriter.Workbook(self.log_file_name)  # In other version we can seperate the workbook class
                worksheet = workbook.add_worksheet("Logs")  # create_log_file should take a instance and use it
                worksheet.set_column(0, 3, 25)
                worksheet.set_column(3, 5, 35)
                worksheet.write("A1", "İşlem Kimliği")
                worksheet.write("B1", "İşlem Durumu")
                worksheet.write("C1", "Tarih")
                worksheet.write("D1", "Başlangıç Zamanı (Saat)")
                worksheet.write("E1", "Bitiş Zamanı (Saat)")
                worksheet.write("F1", "Detaylar")
                workbook.close()
                os.chdir(self.current_dir)
                print("Log dosyası '{0}' dizinine başarılı bir şekilde oluşturuldu" \
                      "...".format(self.log_dir_path))
                print("---" * 37)
            else:
                print("Oluşturulmak istenen log dosyası '{0}' dizininde bulunmaktadır...".format(self.log_dir_path))
                print("---" * 37)
        except FileNotFoundError as error:
            print("Log dosyaları için belirtilen hedef dizin, sistem tarafından oluşturulamadı.")
            print(str(error)[13:])
            print("---" * 37)

    def insert_to_log(self, process_name, process_statu, date, start_time, finish_time, details):

        try:
            open_excel = openpyxl.load_workbook(self.log_file_path) # Excel dosyasını okumak için.
        except FileNotFoundError as error:
            print(error)
            print("Xlsx türündeki log doyası bulunamamıştır.")
        current_cell = None

        try:
            logs_sheet = open_excel.get_sheet_by_name("Logs")
        except KeyError as error:
            print("Xlsx dosyası içerisinde istenilen sheet(çalışma sayfası) bulunamamaktadır.")
        current_cell_number = 2
        logs_sheet = open_excel.active
        for cell_number in range(1, 65536):
            if logs_sheet.cell(row=cell_number, column=1).value == None:
                current_cell_number = cell_number
                break
        # print(current_cell_number)
        logs_sheet["A" + str(current_cell_number)] = process_name
        logs_sheet["B" + str(current_cell_number)] = process_statu
        logs_sheet["C" + str(current_cell_number)] = date
        logs_sheet["D" + str(current_cell_number)] = start_time
        logs_sheet["E" + str(current_cell_number)] = finish_time
        logs_sheet["F" + str(current_cell_number)] = details

        open_excel.save(self.log_file_path)




# Log = LogFile()
# Log.create_log_directory()
# Log.create_log_xlsx_file()

# Log.insert_to_log("Ping taramsı", "12/03/2020", "13:00", "13:50", "Ping başarılıdır.", 2)
